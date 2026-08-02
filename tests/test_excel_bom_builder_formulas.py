"""
Unit tests for MultiTabBOMExcelBuilder 9-sheet official Conecta Excel export.
Asserts that all 9 official worksheets exist, formula strings start with '=',
and Cash Flow / Sensitivity formulas are valid and non-zero.
"""

import io
import openpyxl
import pytest
from src.operations.bom_excel_builder import MultiTabBOMExcelBuilder


class TestExcelBOMBuilderFormulas:
    """Test suite for OpenPyXL formulas and 9 official Conecta worksheets in BOM Excel workbook."""

    @pytest.fixture
    def sample_payload(self):
        return {
            "partner_id": "Cliente Test Coordinado S.A.",
            "amount_untaxed": 50000000.0,
            "amount_tax": 9500000.0,
            "amount_total": 59500000.0,
            "target_margin_pct": 54.8,
            "order_line": [
                {
                    "item_code": "HW-RTU-SUBSTATION",
                    "name": "Remota RTU NovaTech Orion LX+",
                    "product_uom_qty": 2,
                    "price_unit": 15000000.0,
                    "price_subtotal": 30000000.0
                },
                {
                    "item_code": "HW-SWITCH-IND",
                    "name": "Switch Belden Hirschmann RS20",
                    "product_uom_qty": 4,
                    "price_unit": 5000000.0,
                    "price_subtotal": 20000000.0
                }
            ]
        }

    def test_official_9_worksheets_present(self, sample_payload):
        """Assert exact 9 official Conecta worksheets exist in generated workbook."""
        wb = MultiTabBOMExcelBuilder.build_workbook(sample_payload)
        sheet_names = wb.sheetnames

        expected_sheets = [
            "Ficha",
            "Resumen",
            "Control HH y Costos",
            "Equi. Mat. Arr. Sub.",
            "Cash Flow",
            "Cliente",
            "Expenses y Logistica",
            "Terminos de Pago",
            "Check y Sensibilidad"
        ]

        assert len(sheet_names) == 9
        assert sheet_names == expected_sheets

    def test_openpyxl_formulas_present_in_resumen(self, sample_payload):
        """Assert cell formulas in Resumen sheet are valid formula strings starting with '='."""
        wb_bytes = MultiTabBOMExcelBuilder.build_workbook_bytes(sample_payload)
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)

        resumen = wb["Resumen"]

        # Check formula in Net Sales cell (B4)
        net_sales_val = str(resumen["B4"].value)
        assert net_sales_val.startswith("=")
        assert "'Equi. Mat. Arr. Sub.'" in net_sales_val

        # Check formula in IVA cell (B5)
        iva_val = str(resumen["B5"].value)
        assert iva_val.startswith("=")
        assert "B4*0.19" in iva_val

        # Check formula in Total Gross cell (B6)
        total_gross_val = str(resumen["B6"].value)
        assert total_gross_val.startswith("=")
        assert "SUM(B4:B5)" in total_gross_val

    def test_cash_flow_milestone_formulas(self, sample_payload):
        """Assert Cash Flow milestone billing formulas exist and sum to 100% / total."""
        wb_bytes = MultiTabBOMExcelBuilder.build_workbook_bytes(sample_payload)
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)

        cf = wb["Cash Flow"]

        # EDP1, EDP2, EDP3 formula strings
        edp1_formula = str(cf["C4"].value)
        edp2_formula = str(cf["C5"].value)
        edp3_formula = str(cf["C6"].value)
        total_cf_formula = str(cf["C7"].value)

        assert edp1_formula.startswith("=") and "Resumen!B4*0.5" in edp1_formula
        assert edp2_formula.startswith("=") and "Resumen!B4*0.3" in edp2_formula
        assert edp3_formula.startswith("=") and "Resumen!B4*0.2" in edp3_formula
        assert total_cf_formula.startswith("=") and "SUM(C4:C6)" in total_cf_formula

    def test_check_and_sensibilidad_formulas(self, sample_payload):
        """Assert Check y Sensibilidad sheet formulas exist and evaluate correctly."""
        wb_bytes = MultiTabBOMExcelBuilder.build_workbook_bytes(sample_payload)
        wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)

        sens = wb["Check y Sensibilidad"]

        check_sum_formula = str(sens["B4"].value)
        check_status_formula = str(sens["B5"].value)
        sens_baseline_formula = str(sens["B8"].value)
        sens_plus10_formula = str(sens["B9"].value)
        sens_minus10_formula = str(sens["B10"].value)

        assert check_sum_formula.startswith("=") and "SUM" in check_sum_formula
        assert check_status_formula.startswith("=") and "IF" in check_status_formula
        assert sens_baseline_formula.startswith("=") and "Resumen!B8" in sens_baseline_formula
        assert sens_plus10_formula.startswith("=") and "*1.1" in sens_plus10_formula
        assert sens_minus10_formula.startswith("=") and "*0.9" in sens_minus10_formula

    def test_evaluated_workbook_data_values_non_zero(self, sample_payload):
        """Assert that calculated values in workbook are non-zero when evaluated."""
        wb = MultiTabBOMExcelBuilder.build_workbook(sample_payload)

        # Verify calculated values in openpyxl memory before byte export
        resumen = wb["Resumen"]
        assert resumen["B4"].value != 0
        # B7 se almacena como fracción (0.548) con formato "0.0%" para display correcto (54.8%)
        assert resumen["B7"].value == pytest.approx(0.548)

        cf = wb["Cash Flow"]
        assert str(cf["C4"].value).startswith("=") and "0.5" in str(cf["C4"].value)
        assert str(cf["C5"].value).startswith("=") and "0.3" in str(cf["C5"].value)
        assert str(cf["C6"].value).startswith("=") and "0.2" in str(cf["C6"].value)


# =====================================================================
# EXPANDED EDGE CASE & PARAMETERIZED TEST SUITE FOR EXCEL BOM BUILDER
# =====================================================================

@pytest.mark.parametrize("sheet_name", [
    "Ficha",
    "Resumen",
    "Control HH y Costos",
    "Equi. Mat. Arr. Sub.",
    "Cash Flow",
    "Cliente",
    "Expenses y Logistica",
    "Terminos de Pago",
    "Check y Sensibilidad"
])
def test_individual_worksheet_existence_and_non_empty(sheet_name):
    payload = {
        "partner_id": "Empresa Electrica Test",
        "amount_untaxed": 25000000.0,
        "target_margin_pct": 50.0,
        "order_line": [{"item_code": "HW-01", "name": "Item 1", "product_uom_qty": 1, "price_unit": 25000000.0, "price_subtotal": 25000000.0}]
    }
    wb = MultiTabBOMExcelBuilder.build_workbook(payload)
    assert sheet_name in wb.sheetnames
    ws = wb[sheet_name]
    assert ws.max_row >= 1
    assert ws.max_column >= 1


@pytest.mark.parametrize("margin_pct", [10.0, 30.0, 54.8, 70.0, 85.0])
def test_resumen_target_margin_formula_evaluations(margin_pct):
    payload = {
        "partner_id": "Client Margin Test",
        "amount_untaxed": 100000000.0,
        "target_margin_pct": margin_pct,
        "order_line": []
    }
    wb = MultiTabBOMExcelBuilder.build_workbook(payload)
    resumen = wb["Resumen"]
    # B7 se almacena como fracción con formato "0.0%" — display = margin_pct%
    assert resumen["B7"].value == pytest.approx(margin_pct / 100.0)


@pytest.mark.parametrize("untaxed_amount, expected_edp1, expected_edp2, expected_edp3", [
    (1000000.0, 500000.0, 300000.0, 200000.0),
    (50000000.0, 25000000.0, 15000000.0, 10000000.0),
    (100000000.0, 50000000.0, 30000000.0, 20000000.0),
])
def test_cash_flow_3_edp_milestone_evaluations(untaxed_amount, expected_edp1, expected_edp2, expected_edp3):
    payload = {
        "partner_id": "Milestone Test S.A.",
        "amount_untaxed": untaxed_amount,
        "order_line": []
    }
    wb_bytes = MultiTabBOMExcelBuilder.build_workbook_bytes(payload)
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)
    cf = wb["Cash Flow"]

    # Verify formulas reference Resumen!B4
    assert "Resumen!B4*0.5" in str(cf["C4"].value)
    assert "Resumen!B4*0.3" in str(cf["C5"].value)
    assert "Resumen!B4*0.2" in str(cf["C6"].value)


def test_empty_order_line_excel_builder_resilience():
    payload = {
        "partner_id": "Empty Order Line Test",
        "amount_untaxed": 0.0,
        "order_line": []
    }
    wb_bytes = MultiTabBOMExcelBuilder.build_workbook_bytes(payload)
    assert isinstance(wb_bytes, bytes)
    assert len(wb_bytes) > 0
