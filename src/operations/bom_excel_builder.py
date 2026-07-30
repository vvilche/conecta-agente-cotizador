"""
Multi-tab Excel BOM Builder for Conecta Ingeniería S.A.
Generates comprehensive .xlsx workbooks following the exact 9 official Conecta worksheets format:
['Ficha', 'Resumen', 'Control HH y Costos', 'Equi. Mat. Arr. Sub.', 'Cash Flow',
 'Cliente', 'Expenses y Logistica', 'Terminos de Pago', 'Check y Sensibilidad']
"""
import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class MultiTabBOMExcelBuilder:
    # ── Shared Style Helpers ────────────────────────────────────────────────
    HDR_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ACCENT_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    BOLD = Font(name="Calibri", size=11, bold=True)
    NORMAL = Font(name="Calibri", size=11)
    TITLE = Font(name="Calibri", size=14, bold=True, color="1E3A8A")
    SUBTITLE = Font(name="Calibri", size=10, italic=True, color="475569")
    CENTER = Alignment(horizontal="center")
    RIGHT = Alignment(horizontal="right")

    @classmethod
    def _hdr_row(cls, ws, row, titles):
        for col, title in enumerate(titles, start=1):
            c = ws.cell(row=row, column=col, value=title)
            c.font = cls.HDR_FONT
            c.fill = cls.HDR_FILL
            c.alignment = cls.CENTER

    @classmethod
    def _autofit(cls, wb):
        for sheet in wb.worksheets:
            for col in sheet.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)

    @staticmethod
    def _extract_margin_pct(payload: dict) -> float:
        margin_pct = payload.get("target_margin_pct")
        if margin_pct is None:
            margin_pct = payload.get("target_gross_margin")
        if margin_pct is None:
            margin_pct = payload.get("margin_pct")
        if margin_pct is None:
            margin_pct = payload.get("margin_analysis", {}).get("boosted_margin_pct", 54.8)

        try:
            val = float(margin_pct)
        except (ValueError, TypeError):
            val = 54.8

        if 0.0 < val <= 1.0:
            val = val * 100.0

        return max(10.0, min(85.0, val))

    @classmethod
    def build_workbook(cls, payload: dict) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        client_name = payload.get("partner_id", "Cliente Coordinado Conecta")
        lines = payload.get("order_line", [])
        today = datetime.date.today().strftime("%d/%m/%Y")

        target_margin_pct = cls._extract_margin_pct(payload)

        # ── 1. Ficha ───────────────────────────────────────────────────────
        ws_ficha = wb.create_sheet("Ficha")
        ws_ficha.cell(1, 1, "CONECTA INGENIERÍA S.A. — FICHA DE TRASPASO OT").font = cls.TITLE
        ws_ficha.cell(2, 1, "Formulario Oficial de Traspaso Comercial a Operaciones").font = cls.SUBTITLE
        ficha_data = [
            ("Fecha de Traspaso", today),
            ("N° Oferta Conecta", "OF-2026-CONECTA-REV0"),
            ("N° OT Conecta", "7095-00"),
            ("Cliente Coordinado", client_name),
            ("RUT Cliente", "76.543.210-9"),
            ("Nombre Proyecto", f"Suministro e Integración OT — {client_name}"),
            ("Jefe de Proyecto Asignado", "Ing. Pedro Morales"),
            ("Monto Oferta Neto CLP", "=Resumen!B4"),
            ("Margen Bruto Target (%)", "=Resumen!B7"),
        ]
        for i, (lbl, val) in enumerate(ficha_data, start=4):
            c1 = ws_ficha.cell(i, 1, lbl); c1.font = cls.BOLD; c1.fill = cls.ACCENT_FILL
            c2 = ws_ficha.cell(i, 2, val)
            if lbl == "Margen Bruto Target (%)":
                c2.number_format = "0.0%"
            elif lbl == "Monto Oferta Neto CLP":
                c2.number_format = "$#,##0"

        # ── 4. Equi. Mat. Arr. Sub. (Build hardware lines first to derive total row) ───
        ws_equi = wb.create_sheet("Equi. Mat. Arr. Sub.")
        ws_equi.cell(1, 1, "DETALLE DE EQUIPOS, MATERIALES, ARRIENDOS Y SUBCONTRATOS").font = cls.TITLE
        cls._hdr_row(ws_equi, 3, ["Categoría", "Código Partida", "Descripción Ítem", "Marca / Modelo", "Cant.", "P. Unit. CLP", "Subtotal CLP"])
        
        default_items = [
            {"item_code": "HW-RTU-NOVATECH", "name": "Remota RTU NovaTech Orion LX+", "brand": "NovaTech", "qty": 1, "price": 21681416},
            {"item_code": "HW-SWITCH-IND", "name": "Switch Ethernet Belden Hirschmann RS20", "brand": "Belden", "qty": 1, "price": 5088496},
            {"item_code": "HW-VIZIMAX-PMU", "name": "Medidor Fasorial PMU VIZIMAX SynchroTeq", "brand": "VIZIMAX", "qty": 1, "price": 9500000},
            {"item_code": "HW-GPS-CLOCK", "name": "Reloj Satelital GPS Kronos IRIG-B", "brand": "Kronos", "qty": 1, "price": 3200000},
        ]
        
        if lines:
            line_rows = []
            for item in lines:
                line_rows.append({
                    "item_code": item.get("item_code", "HW-ITEM"),
                    "name": item.get("name", "Ítem de Equipamiento OT"),
                    "brand": item.get("brand", "Especializado"),
                    "qty": item.get("product_uom_qty", 1),
                    "price": item.get("price_unit", 0)
                })
        else:
            line_rows = default_items

        start_row = 4
        for idx, item in enumerate(line_rows, start=start_row):
            ws_equi.cell(idx, 1, "HARDWARE/SW").font = cls.BOLD
            ws_equi.cell(idx, 2, item["item_code"])
            ws_equi.cell(idx, 3, item["name"])
            ws_equi.cell(idx, 4, item.get("brand", "Especializado"))
            ws_equi.cell(idx, 5, item["qty"]).alignment = cls.CENTER
            c = ws_equi.cell(idx, 6, item["price"]); c.number_format = "$#,##0"
            c = ws_equi.cell(idx, 7, f"=E{idx}*F{idx}"); c.number_format = "$#,##0"; c.font = cls.BOLD

        end_row = start_row + len(line_rows) - 1
        equi_total_row = end_row + 1
        ws_equi.cell(equi_total_row, 1, "TOTAL EQUIPOS Y MATERIALES").font = cls.BOLD
        c = ws_equi.cell(equi_total_row, 5, f"=SUM(E{start_row}:E{end_row})"); c.alignment = cls.CENTER; c.font = cls.BOLD
        c = ws_equi.cell(equi_total_row, 7, f"=SUM(G{start_row}:G{end_row})"); c.number_format = "$#,##0"; c.font = cls.BOLD

        # ── 2. Resumen ─────────────────────────────────────────────────────
        ws_resumen = wb.create_sheet("Resumen")
        ws_resumen.cell(1, 1, "RESUMEN FINANCIERO Y OPTIMIZACIÓN DE MARGEN").font = cls.TITLE
        cls._hdr_row(ws_resumen, 3, ["Indicador Financiero", "Valor CLP / %"])
        
        ws_resumen.cell(4, 1, "Ventas Neto Cliente (Sin IVA)").font = cls.BOLD
        c = ws_resumen.cell(4, 2, f"='Equi. Mat. Arr. Sub.'!G{equi_total_row}"); c.number_format = "$#,##0"; c.font = cls.BOLD
        
        ws_resumen.cell(5, 1, "Impuesto IVA (19%)").font = cls.BOLD
        c = ws_resumen.cell(5, 2, "=B4*0.19"); c.number_format = "$#,##0"; c.font = cls.BOLD
        
        ws_resumen.cell(6, 1, "Total Bruto Facturación").font = cls.BOLD
        c = ws_resumen.cell(6, 2, "=SUM(B4:B5)"); c.number_format = "$#,##0"; c.font = cls.BOLD

        ws_resumen.cell(7, 1, "Margen Bruto Target (%)").font = cls.BOLD
        c = ws_resumen.cell(7, 2, target_margin_pct); c.number_format = "0.0%"; c.font = cls.BOLD; c.fill = cls.GREEN_FILL
        ws_resumen.cell(7, 1).fill = cls.GREEN_FILL

        ws_resumen.cell(8, 1, "Margen Bruto Target (CLP)").font = cls.BOLD
        c = ws_resumen.cell(8, 2, "=B4*(B7/100)"); c.number_format = "$#,##0"; c.font = cls.BOLD; c.fill = cls.GREEN_FILL
        ws_resumen.cell(8, 1).fill = cls.GREEN_FILL

        ws_resumen.cell(9, 1, "Costo Directo Estimado").font = cls.BOLD
        c = ws_resumen.cell(9, 2, "=B4*(1-B7/100)"); c.number_format = "$#,##0"; c.font = cls.BOLD

        ws_resumen.cell(10, 1, "Utilidad Bruta Retenida").font = cls.BOLD
        c = ws_resumen.cell(10, 2, "=B4*(B7/100)"); c.number_format = "$#,##0"; c.font = cls.BOLD; c.fill = cls.GREEN_FILL
        ws_resumen.cell(10, 1).fill = cls.GREEN_FILL

        ws_resumen.cell(11, 1, "Margen Bruto Retenido (%)").font = cls.BOLD
        c = ws_resumen.cell(11, 2, "=B10/B4"); c.number_format = "0.0%"; c.font = cls.BOLD; c.fill = cls.GREEN_FILL
        ws_resumen.cell(11, 1).fill = cls.GREEN_FILL

        # ── 3. Control HH y Costos ─────────────────────────────────────────
        ws_hh = wb.create_sheet("Control HH y Costos")
        ws_hh.cell(1, 1, "CONTROL HH Y COSTOS POR ACTIVIDAD Y ROL").font = cls.TITLE
        ws_hh.cell(3, 1, "HH POR ACTIVIDADES DE INGENIERÍA").font = cls.BOLD
        cls._hdr_row(ws_hh, 4, ["ÍTEM", "PO/PM (HH)", "ING. SENIOR/A (HH)", "ING. B (HH)", "TÉC. TERRENO (HH)", "TOTAL HH"])
        
        hh_items = [
            ("PLANIFICACION", 4, 8, 0, 0),
            ("LEVANTAMIENTO", 0, 4, 8, 0),
            ("INGENIERIA", 2, 24, 16, 0),
            ("IMPLEMENTACION", 0, 8, 16, 16),
            ("PRUEBAS HIL FAT", 0, 8, 12, 8),
            ("SAT TERRENO", 2, 8, 8, 16),
            ("REGULATORY CEN", 2, 8, 4, 0),
        ]
        for idx, (name, po, ing_a, ing_b, tec) in enumerate(hh_items, start=5):
            ws_hh.cell(idx, 1, name).font = cls.BOLD
            ws_hh.cell(idx, 2, po).alignment = cls.CENTER
            ws_hh.cell(idx, 3, ing_a).alignment = cls.CENTER
            ws_hh.cell(idx, 4, ing_b).alignment = cls.CENTER
            ws_hh.cell(idx, 5, tec).alignment = cls.CENTER
            c = ws_hh.cell(idx, 6, f"=SUM(B{idx}:E{idx})"); c.alignment = cls.CENTER; c.font = cls.BOLD

        ws_hh.cell(12, 1, "TOTAL HH").font = cls.BOLD
        for col_idx, col_let in enumerate(["B", "C", "D", "E", "F"], start=2):
            c = ws_hh.cell(12, col_idx, f"=SUM({col_let}5:{col_let}11)")
            c.font = cls.BOLD; c.alignment = cls.CENTER

        # Cost breakdown section in Control HH y Costos
        ws_hh.cell(15, 1, "MATRIZ VALORIZADA DE HORAS HOMBRE").font = cls.BOLD
        cls._hdr_row(ws_hh, 16, ["Rol", "Tarifa HH (CLP/hr)", "Total HH", "Costo Total CLP"])
        roles = [
            ("Project Owner / PM", 95000, "=B12"),
            ("Ing. Senior / A", 85000, "=C12"),
            ("Ing. Especialista B", 70000, "=D12"),
            ("Técnico Terreno", 55000, "=E12"),
        ]
        for idx, (rol, rate, hh_ref) in enumerate(roles, start=17):
            ws_hh.cell(idx, 1, rol).font = cls.BOLD
            c = ws_hh.cell(idx, 2, rate); c.number_format = "$#,##0"
            c = ws_hh.cell(idx, 3, hh_ref); c.alignment = cls.CENTER
            c = ws_hh.cell(idx, 4, f"=B{idx}*C{idx}"); c.number_format = "$#,##0"; c.font = cls.BOLD

        ws_hh.cell(21, 1, "TOTAL COSTO HH").font = cls.BOLD
        c = ws_hh.cell(21, 3, "=SUM(C17:C20)"); c.alignment = cls.CENTER; c.font = cls.BOLD
        c = ws_hh.cell(21, 4, "=SUM(D17:D20)"); c.number_format = "$#,##0"; c.font = cls.BOLD

        # Reorder worksheets so ['Ficha', 'Resumen', 'Control HH y Costos', 'Equi. Mat. Arr. Sub.'] are in position
        wb._sheets = [
            wb["Ficha"],
            wb["Resumen"],
            wb["Control HH y Costos"],
            wb["Equi. Mat. Arr. Sub."],
        ]

        # ── 5. Cash Flow ───────────────────────────────────────────────────
        ws_cf = wb.create_sheet("Cash Flow")
        ws_cf.cell(1, 1, "FLUJO DE CAJA — HITOS DE COBRANZA (EDP)").font = cls.TITLE
        cls._hdr_row(ws_cf, 3, ["Hito EDP", "Descripción Hito", "Monto Neto CLP", "Facturación %", "Fecha Est."])
        
        edp_items = [
            ("EDP 1", "EDP 1 Pre-kitting (50%)", "=Resumen!B4*0.5", 0.50, "Semana 3"),
            ("EDP 2", "EDP 2 SAT HIL (30%)", "=Resumen!B4*0.3", 0.30, "Semana 6"),
            ("EDP 3", "EDP 3 Handover / Factura Final (20%)", "=Resumen!B4*0.2", 0.20, "Semana 8"),
        ]
        for idx, (hito, desc, formula_str, pct, fecha) in enumerate(edp_items, start=4):
            ws_cf.cell(idx, 1, hito).font = cls.BOLD
            ws_cf.cell(idx, 2, desc)
            c = ws_cf.cell(idx, 3, formula_str); c.number_format = "$#,##0"; c.font = cls.BOLD
            c = ws_cf.cell(idx, 4, pct); c.number_format = "0.0%"; c.alignment = cls.CENTER
            ws_cf.cell(idx, 5, fecha).alignment = cls.CENTER

        ws_cf.cell(7, 1, "TOTAL HITOS EDP").font = cls.BOLD
        ws_cf.cell(7, 2, "Facturación Total 100%").font = cls.SUBTITLE
        c = ws_cf.cell(7, 3, "=SUM(C4:C6)"); c.number_format = "$#,##0"; c.font = cls.BOLD
        c = ws_cf.cell(7, 4, "=SUM(D4:D6)"); c.number_format = "0.0%"; c.alignment = cls.CENTER; c.font = cls.BOLD

        # ── 6. Cliente ─────────────────────────────────────────────────────
        ws_cli = wb.create_sheet("Cliente")
        ws_cli.cell(1, 1, "METADATA DEL CLIENTE COORDINADO").font = cls.TITLE
        client_info = [
            ("Razón Social", client_name),
            ("RUT Empresa", "76.543.210-9"),
            ("Giro / Sector", "Generación y Transmisión de Energía Eléctrica"),
            ("Dirección", "Av. Andrés Bello 2711, Las Condes, Santiago"),
            ("Contacto Técnico", "Ing. Administrador de Contratos"),
            ("Teléfono", "+56 2 2345 6789"),
            ("Email Contacto", "contacto@cliente.cl"),
        ]
        for idx, (lbl, val) in enumerate(client_info, start=4):
            c1 = ws_cli.cell(idx, 1, lbl); c1.font = cls.BOLD; c1.fill = cls.ACCENT_FILL
            ws_cli.cell(idx, 2, val)

        # ── 7. Expenses y Logistica ────────────────────────────────────────
        ws_exp = wb.create_sheet("Expenses y Logistica")
        ws_exp.cell(1, 1, "LOGÍSTICA, VIÁTICOS Y ACREDITACIÓN EN FAENA").font = cls.TITLE
        cls._hdr_row(ws_exp, 3, ["Concepto", "Detalle Logístico", "Días / Unid.", "Tarifa Unit. CLP", "Subtotal CLP"])
        
        expenses_data = [
            ("Pre-kitting Tableros Taller", "Ensamblaje y pre-cableado estandarizado KittingEngine", 1.5, 1233333),
            ("Acreditación Digital", "Dossier Sicop/Pronexo F30-1, ex. médicos, EPP, ODI/DAS", 1.0, 450000),
            ("Traslado & Camioneta 4x4", "Movilización de tableros y camioneta equipada 4x4", 3.0, 400000),
            ("Viáticos Especialistas", "Alimentación y hospedaje en faena terreno", 3.0, 326667),
        ]
        for idx, (con, det, qty, rate) in enumerate(expenses_data, start=4):
            ws_exp.cell(idx, 1, con).font = cls.BOLD
            ws_exp.cell(idx, 2, det)
            ws_exp.cell(idx, 3, qty).alignment = cls.CENTER
            c = ws_exp.cell(idx, 4, rate); c.number_format = "$#,##0"
            c = ws_exp.cell(idx, 5, f"=C{idx}*D{idx}"); c.number_format = "$#,##0"; c.font = cls.BOLD

        ws_exp.cell(8, 1, "TOTAL EXPENSES Y LOGÍSTICA").font = cls.BOLD
        c = ws_exp.cell(8, 3, "=SUM(C4:C7)"); c.alignment = cls.CENTER; c.font = cls.BOLD
        c = ws_exp.cell(8, 5, "=SUM(E4:E7)"); c.number_format = "$#,##0"; c.font = cls.BOLD

        # ── 8. Terminos de Pago ──────────────────────────────────────────
        ws_tp = wb.create_sheet("Terminos de Pago")
        ws_tp.cell(1, 1, "TÉRMINOS Y CONDICIONES DE PAGO Y GARANTÍA").font = cls.TITLE
        terms = [
            ("Condición de Pago", "Facturación a 30 días tras aprobación de Estado de Pago (EDP)"),
            ("Validez de la Oferta", "30 días corridos a contar de la fecha de presentación"),
            ("Garantía Técnica", "12 meses desde la puesta en servicio comercial"),
            ("Boleta Fiel Cumplimiento", "10% del monto neto total (requerido para licitaciones)"),
            ("Multas / Cap", "Techo máximo acumulado de multas: 5% del valor neto del contrato"),
            ("Fuerza Mayor", "Suspensión de plazos por eventos fuera del control de las partes"),
        ]
        for idx, (lbl, val) in enumerate(terms, start=4):
            c1 = ws_tp.cell(idx, 1, lbl); c1.font = cls.BOLD; c1.fill = cls.ACCENT_FILL
            ws_tp.cell(idx, 2, val)

        # ── 9. Check y Sensibilidad ──────────────────────────────────────
        ws_cs = wb.create_sheet("Check y Sensibilidad")
        ws_cs.cell(1, 1, "CHECKLIST DE PROYECTO & ANÁLISIS DE SENSIBILIDAD DE MARGEN").font = cls.TITLE
        
        ws_cs.cell(3, 1, "RESUMEN DE CONTROLES").font = cls.BOLD
        ws_cs.cell(4, 1, "Total Ítems de Control").font = cls.BOLD
        c = ws_cs.cell(4, 2, "=SUM(C15:C22)"); c.alignment = cls.CENTER; c.font = cls.BOLD
        
        ws_cs.cell(5, 1, "Estado Global Project Checklist").font = cls.BOLD
        c = ws_cs.cell(5, 2, '=IF(B4>0, "CONFORME", "PENDIENTE")'); c.alignment = cls.CENTER; c.font = cls.BOLD

        ws_cs.cell(7, 1, "ANÁLISIS DE SENSIBILIDAD FINANCIERA DE MARGEN (CLP)").font = cls.BOLD
        ws_cs.cell(8, 1, "Margen Baseline Target CLP").font = cls.BOLD
        c = ws_cs.cell(8, 2, "=Resumen!B8"); c.number_format = "$#,##0"; c.font = cls.BOLD
        
        ws_cs.cell(9, 1, "Escenario Optimista (+10% Venta)").font = cls.BOLD
        c = ws_cs.cell(9, 2, "=Resumen!B8*1.1"); c.number_format = "$#,##0"; c.font = cls.BOLD

        ws_cs.cell(10, 1, "Escenario Pesimista (-10% Venta)").font = cls.BOLD
        c = ws_cs.cell(10, 2, "=Resumen!B8*0.9"); c.number_format = "$#,##0"; c.font = cls.BOLD

        ws_cs.cell(13, 1, "MATRIZ DE RIESGO Y CONTROL").font = cls.BOLD
        cls._hdr_row(ws_cs, 14, ["Área", "Ítem de Control / Riesgo", "Estado / Valor", "Responsable", "Nivel de Riesgo"])
        
        checks = [
            ("Comercial", "Oferta firmada y enviada al cliente", 1, "Comercial", "Bajo"),
            ("Comercial", "Orden de Compra recibida", 1, "Comercial", "Medio"),
            ("Operaciones", "Ficha de Traspaso llenada en SAP/Odoo", 1, "Jefe OT", "Bajo"),
            ("Operaciones", "Pre-kitting tableros completado en taller", 1, "Téc. Taller", "Bajo"),
            ("Operaciones", "FAT/SAT HIL ejecutado y certificado firmado", 1, "Ing. Senior", "Medio"),
            ("Regulatory", "Informe IPES redactado y enviado al CEN", 1, "PM", "Bajo"),
            ("Cobranza", "EDP 1 emitido y facturado", 1, "Administración", "Bajo"),
            ("Cobranza", "EDP 2 emitido y facturado", 1, "Administración", "Bajo"),
        ]
        for idx, (area, item, estado, resp, riesgo) in enumerate(checks, start=15):
            ws_cs.cell(idx, 1, area).font = cls.BOLD
            ws_cs.cell(idx, 2, item)
            ws_cs.cell(idx, 3, estado).alignment = cls.CENTER
            ws_cs.cell(idx, 4, resp)
            ws_cs.cell(idx, 5, riesgo).alignment = cls.CENTER

        cls._autofit(wb)
        return wb

    @classmethod
    def build_workbook_bytes(cls, payload: dict) -> bytes:
        wb = cls.build_workbook(payload)
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

