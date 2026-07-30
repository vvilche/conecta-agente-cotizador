#!/usr/bin/env python3
"""
Unique Project Auditor & Deduplication Engine.
Filters out duplicate file revisions (Rev 0, Rev A, Rev B, Rev C vs Rev D) to present the EXACT, 1-to-1 list of unique projects and their final values for 2025 and 2026.
"""

import sqlite3
import openpyxl
import pandas as pd
from pathlib import Path
import re

def audit_unique_projects():
    dir_2025 = Path("2025")
    
    # 1. Group files by top-level project folder in 2025
    project_folders = {}
    
    if dir_2025.exists():
        # Find directories inside 2025 that look like project folders
        for p in dir_2025.rglob("*"):
            if p.is_dir() and re.match(r"^\d{5,6}\s*-", p.name):
                project_folders[p.name] = p

    print("==========================================================================")
    print("🔎 REVISIÓN DE PROYECTOS ÚNICOS 2025 (ELIMINANDO DUPLICACIONES DE REVISIÓN)")
    print("==========================================================================")
    print(f"• Total Carpetas Únicas de Proyecto Identificadas en 2025: {len(project_folders)}")

    unique_projects = []

    for p_name, p_path in sorted(project_folders.items()):
        # Find final calculation file in 3 Calculo de la Oferta or root
        xl_files = list(p_path.rglob("*.xlsm")) + list(p_path.rglob("*.xlsx"))
        
        # Filter out ~$, temporary, or duplicate files
        xl_files = [f for f in xl_files if not f.name.startswith("~$") and "obsoleto" not in str(f).lower()]

        if not xl_files:
            continue

        # Sort files to pick the latest revision (Rev D > Rev C > Rev 0)
        def rev_key(f):
            name = f.name.lower()
            if "rev d" in name: return 4
            if "rev c" in name: return 3
            if "rev b" in name: return 2
            if "rev a" in name: return 1
            if "rev 0" in name: return 0
            return -1

        xl_files.sort(key=rev_key, reverse=True)
        final_file = xl_files[0]

        # Extract client name
        client = "Otros Clientes"
        if "transelec" in p_name.lower(): client = "Transelec"
        elif "chilquinta" in p_name.lower(): client = "Chilquinta"
        elif "aes" in p_name.lower(): client = "AES Andes"
        elif "colbun" in p_name.lower(): client = "Colbún"
        elif "saesa" in p_name.lower(): client = "SAESA"
        elif "cge" in p_name.lower(): client = "CGE"
        elif "enel" in p_name.lower(): client = "Enel"

        # Attempt to read exact value
        val_clp = 0.0
        val_uf = 0.0
        try:
            wb = openpyxl.load_workbook(final_file, read_only=True, data_only=True)
            if "Resumen" in wb.sheetnames:
                sheet = wb["Resumen"]
                for row in sheet.iter_rows(values_only=True):
                    if len(row) > 1 and row[1]:
                        txt = str(row[1]).strip().upper()
                        if "TOTAL GLOBAL NETO" in txt or "TOTAL GLOBAL BRUTO" in txt:
                            v = row[4] if len(row) > 4 and row[4] else (row[7] if len(row) > 7 else None)
                            if isinstance(v, (int, float)) and v > 0:
                                if v < 100000:
                                    val_uf = float(v)
                                    val_clp = val_uf * 38377.09
                                else:
                                    val_clp = float(v)
                                break
        except Exception:
            pass

        unique_projects.append({
            "folder": p_name,
            "client": client,
            "final_file": final_file.name,
            "val_uf": val_uf,
            "val_clp": val_clp
        })

    print(f"\n{'N°':<3} | {'PROYECTO':<45} | {'CLIENTE':<12} | {'ARCHIVO FINAL':<40} | {'MONTO UF':<10} | {'MONTO NETO CLP'}")
    print("-" * 140)

    total_2025_clp = 0.0
    for idx, proj in enumerate(unique_projects, 1):
        f_name = proj["folder"][:45]
        cl = proj["client"][:12]
        ff = proj["final_file"][:40]
        uf = proj["val_uf"]
        clp = proj["val_clp"]
        total_2025_clp += clp

        uf_str = f"{uf:,.1f} UF" if uf > 0 else "-"
        clp_str = f"${clp:,.0f} CLP" if clp > 0 else "Por Determinar"

        print(f"{idx:<3} | {f_name:<45} | {cl:<12} | {ff:<40} | {uf_str:>10} | {clp_str:>16}")

    print("-" * 140)
    print(f"TOTAL REAL PROYECTOS ÚNICOS 2025: ${total_2025_clp:,.0f} CLP (~${total_2025_clp/1e6:.2f} Millones de CLP)\n")

if __name__ == "__main__":
    audit_unique_projects()
